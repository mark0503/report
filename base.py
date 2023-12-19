import openpyxl

from utils import logger


class BaseProcessor:

    def __init__(
            self,
            source_data
    ):

        self.count_column: int = source_data['column_count']
        self.percent: int = source_data['percent']
        self.max_count: int = source_data['max_count']
        self.max_amount: int = source_data['max_amount']
        self.column_amount: int = source_data['column_amount']
        self.need_long_id: bool = source_data['need_long_id']
        self.file_name: str = source_data['file_name']
        self.workbook = None
        self.sheet = None
        self.state: dict = {
            'initial_list': [],
            'result_list': []
        }

    async def initial_process(self):
        logger.info(f'Начинаю читать файл {self.file_name}')
        try:
            self.workbook = openpyxl.load_workbook(self.file_name, data_only=True,
                                                   keep_links=False)
            self.sheet = self.workbook.active
        except Exception as err:
            logger.error(f'Ошибка обработки файла {self.file_name}: {err}')
        else:
            logger.success(f'Файл {self.file_name} успешно обработан. Кол-во строк: {self.sheet.max_row}')

    async def save_process(self):
        for element in self.state['result_list']:
            self.sheet.cell(row=element['row'], column=21).value = element['new_user_id']

        print('начинаем сохранять файл')
        self.workbook.save(f'{self.file_name.replace(".xlsx", "")}_patch.xlsx')
