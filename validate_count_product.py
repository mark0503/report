import openpyxl

from utils import logger


class ValidateCountProduct:

    def __init__(self, source_data):

        self.file_name = source_data['file_name']
        self.description_column = source_data['description_column']
        self.save_column = source_data['save_column']
        self.incorrect_keywords_list = source_data['incorrect_keywords_list']
        self.true_keywords_list = source_data['true_keywords_list']
        self.workbook = None
        self.sheet = None

    async def initial_process(self):
        try:
            logger.success(f'Начинаем читать файл {self.file_name}.')
            self.workbook = openpyxl.load_workbook(self.file_name, data_only=True,
                                                   keep_links=False)
            self.sheet = self.workbook.active
        except Exception as err:
            logger.error(f'Ошибка обработки файла {self.file_name}: {err}')
        else:
            logger.success(f'Файл {self.file_name} успешно прочитан. Кол-во строк: {self.sheet.max_row}')

    async def start_validate(self):
        await self.initial_process()
        await self.start_process()
        await self.save_process()
        return True

    async def start_process(self):
        for row in range(2, self.sheet.max_row + 1):
            string_input = self.sheet.cell(row=row, column=self.description_column).value
            if not string_input:
                continue
            count = 0
            is_preview_valid = False
            for i in string_input.split('кол-во - '):
                count_str = ''
                if i.isdigit() and is_preview_valid:
                    count += int(i)
                else:
                    for value in i:
                        if value.isdigit():
                            count_str += value
                            if len(i) == 1 and is_preview_valid:
                                count += int(count_str)
                        else:
                            if count_str and is_preview_valid:
                                count += int(count_str)
                            break
                    is_broke = False
                    for false_string in self.incorrect_keywords_list:
                        if false_string in i.lower():
                            is_broke = True
                            break
                    for true_string in self.true_keywords_list:
                        if true_string not in i.lower():
                            is_broke = True
                            break
                    if is_broke is True:
                        is_preview_valid = False
                    else:
                        is_preview_valid = True
            self.sheet.cell(row=row, column=self.save_column).value = count

    async def save_process(self):
        logger.success(f'Начинаем сохранять файл {self.file_name}.')
        self.workbook.save(f'{self.file_name.replace(".xlsx", "")}_patch.xlsx')
        logger.success(f'Файл {self.file_name} успешно сохранен.')
