import openpyxl

from utils import logger
from utils.generate_ex_id import generate_id_of_type


class SetUserIdAtExID:

    def __init__(self, source_data):

        self.user_ex_id_dict = None
        self.file_name = source_data['file_name']
        self.save_column = source_data['save_column']
        self.ex_id = int(source_data['ex_id'])
        self.workbook = None
        self.id_type = source_data['id_type']
        self.sheet = None

    def initial_process(self):
        try:
            logger.success(f'Начинаем читать файл {self.file_name}.')
            self.workbook = openpyxl.load_workbook(self.file_name, data_only=True,
                                                   keep_links=False)
            self.sheet = self.workbook.active
        except Exception as err:
            logger.error(f'Ошибка обработки файла {self.file_name}: {err}')
        else:
            logger.success(f'Файл {self.file_name} успешно прочитан. Кол-во строк: {self.sheet.max_row}')

    def start_validate(self):
        self.user_ex_id_dict = {}
        self.initial_process()
        self.start_process()
        self.save_process()
        return True

    def start_process(self):
        logger.success(f'Начинаем обрабатывать файл {self.file_name}.')
        for row in range(2, self.sheet.max_row + 1):
            user_ex_id = self.sheet.cell(row=row, column=self.ex_id).value

            if self.user_ex_id_dict.get(user_ex_id):
                uuids_user_id = self.user_ex_id_dict[user_ex_id]
            else:
                uuids_user_id = generate_id_of_type(self.id_type)
                self.user_ex_id_dict[user_ex_id] = uuids_user_id

            self.sheet.cell(row=row, column=self.save_column).value = uuids_user_id

    def save_process(self):
        logger.success(f'Начинаем сохранять файл {self.file_name}.')
        self.workbook.save(f'{self.file_name.replace(".xlsx", "")}_patch.xlsx')
        logger.success(f'Файл {self.file_name} успешно сохранен.')
